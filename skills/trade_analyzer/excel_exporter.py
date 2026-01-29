"""
Excel Exporter - Excel 导出模块

使用 openpyxl 导出交易记录和统计数据到 Excel 文件：
- Sheet 1: 股票交易明细
- Sheet 2: 期权交易明细
- Sheet 3: 统计汇总
"""

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from .statistics import TradeStatistics, StatisticsCalculator
from .trade_matcher import MatchedTrade


class ExcelExporter:
    """Excel 导出器"""

    # 样式定义
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    PROFIT_FILL = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    PROFIT_FONT = Font(color="006100")
    LOSS_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    LOSS_FONT = Font(color="9C0006")

    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 列定义
    TRADE_COLUMNS = [
        ("交易序号", 10),
        ("市场", 8),
        ("交易标的", 25),
        ("买入价", 12),
        ("买入数量", 12),
        ("买入市值", 15),
        ("卖出价", 12),
        ("卖出数量", 12),
        ("卖出市值", 15),
        ("盈亏额", 15),
        ("盈亏率", 12),
        ("买入日期", 12),
        ("卖出日期", 12),
        ("持仓天数", 10),
        ("盈/亏", 8),
        ("滚动胜率", 12),
        ("滚动平均盈利", 15),
        ("滚动平均亏损", 15),
        ("滚动盈亏比", 12),
    ]

    def __init__(self):
        self.wb = Workbook()

    def export(
        self,
        trades: list[MatchedTrade],
        stats: TradeStatistics,
        output_path: Path,
        year: Optional[int] = None,
    ) -> Path:
        """
        导出交易数据到 Excel

        Args:
            trades: 配对后的交易列表
            stats: 统计数据
            output_path: 输出文件路径
            year: 年份（用于文件名）

        Returns:
            输出文件路径
        """
        self.wb = Workbook()

        # 分离股票和期权交易
        stock_trades = [t for t in trades if not t.is_option]
        option_trades = [t for t in trades if t.is_option]

        # Sheet 1: 股票交易明细
        self._create_trade_sheet(stock_trades, "股票交易明细", year)

        # Sheet 2: 期权交易明细
        if option_trades:
            self._create_trade_sheet(
                option_trades, "期权交易明细", year, is_option=True
            )

        # Sheet 3: 统计汇总
        self._create_summary_sheet(stats, year)

        # 删除默认的 Sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # 保存文件
        self.wb.save(output_path)
        return output_path

    def _create_trade_sheet(
        self,
        trades: list[MatchedTrade],
        sheet_name: str,
        year: Optional[int] = None,
        is_option: bool = False,
    ) -> None:
        """创建交易明细表"""
        ws = self.wb.create_sheet(sheet_name)

        # 标题行
        title = f"{year or datetime.now().year}年美港股{'期权' if is_option else '股票'}交易记录"
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=len(self.TRADE_COLUMNS)
        )
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        # 表头
        for col_idx, (col_name, col_width) in enumerate(self.TRADE_COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        # 按卖出日期排序
        sorted_trades = sorted(
            trades,
            key=lambda t: t.sell_date or datetime.min,
            reverse=True,
        )

        # 滚动计算变量
        running_wins = 0
        running_total = 0
        running_profit_sum = Decimal("0")
        running_loss_sum = Decimal("0")
        running_profit_count = 0
        running_loss_count = 0

        # 数据行
        for row_idx, trade in enumerate(sorted_trades, start=3):
            running_total += 1
            if trade.is_profitable:
                running_wins += 1
                running_profit_sum += trade.profit_loss
                running_profit_count += 1
            elif trade.profit_loss < 0:
                running_loss_sum += abs(trade.profit_loss)
                running_loss_count += 1

            # 计算滚动指标
            rolling_win_rate = running_wins / running_total if running_total > 0 else 0
            rolling_avg_profit = (
                running_profit_sum / running_profit_count
                if running_profit_count > 0
                else Decimal("0")
            )
            rolling_avg_loss = (
                running_loss_sum / running_loss_count
                if running_loss_count > 0
                else Decimal("0")
            )
            rolling_pl_ratio = (
                rolling_avg_profit / rolling_avg_loss
                if rolling_avg_loss > 0
                else Decimal("999")
            )

            row_data = [
                running_total,  # 交易序号
                trade.market,
                trade.stock_name or trade.code,
                float(trade.buy_price),
                float(trade.buy_qty),
                float(trade.buy_amount),
                float(trade.sell_price),
                float(trade.sell_qty),
                float(trade.sell_amount),
                float(trade.profit_loss),
                float(trade.profit_loss_ratio),
                trade.buy_date.strftime("%Y-%m-%d") if trade.buy_date else "",
                trade.sell_date.strftime("%Y-%m-%d") if trade.sell_date else "",
                trade.holding_days,
                "盈" if trade.is_profitable else "亏",
                f"{rolling_win_rate:.1%}",
                float(rolling_avg_profit),
                float(rolling_avg_loss),
                float(rolling_pl_ratio) if rolling_pl_ratio < 999 else "∞",
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal="center")

                # 数字格式
                if col_idx in (4, 5, 6, 7, 8, 9, 10, 17, 18):  # 金额列
                    cell.number_format = "#,##0.00"
                elif col_idx == 11:  # 盈亏率
                    cell.number_format = "0.00%"
                elif col_idx == 19 and isinstance(value, float):  # 盈亏比
                    cell.number_format = "0.00"

            # 盈亏高亮
            profit_loss_cell = ws.cell(row=row_idx, column=10)
            status_cell = ws.cell(row=row_idx, column=15)

            if trade.is_profitable:
                profit_loss_cell.fill = self.PROFIT_FILL
                profit_loss_cell.font = self.PROFIT_FONT
                status_cell.fill = self.PROFIT_FILL
                status_cell.font = self.PROFIT_FONT
            elif trade.profit_loss < 0:
                profit_loss_cell.fill = self.LOSS_FILL
                profit_loss_cell.font = self.LOSS_FONT
                status_cell.fill = self.LOSS_FILL
                status_cell.font = self.LOSS_FONT

    def _create_summary_sheet(
        self, stats: TradeStatistics, year: Optional[int] = None
    ) -> None:
        """创建统计汇总表"""
        ws = self.wb.create_sheet("统计汇总")

        # 标题
        title = f"{year or datetime.now().year}年交易统计汇总"
        ws.merge_cells("A1:C1")
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)

        row = 3

        # 整体统计
        row = self._add_section_header(ws, row, "一、整体交易表现")
        summary_data = [
            ("总交易笔数", f"{stats.total_trades}笔", "已配对完成的买卖交易"),
            ("盈利笔数", f"{stats.winning_trades}笔", f"占比{stats.win_rate:.1%}"),
            ("亏损笔数", f"{stats.losing_trades}笔", f"占比{1-stats.win_rate:.1%}"),
            ("胜率", f"{stats.win_rate:.1%}", "盈利交易占总交易的比例"),
            ("盈亏比", f"{float(stats.profit_loss_ratio):.2f}", "平均盈利/平均亏损"),
        ]
        row = self._add_table(ws, row, ["指标", "数值", "说明"], summary_data)

        # 盈亏统计
        row = self._add_section_header(ws, row + 1, "二、盈亏统计")
        profit_data = [
            ("总盈利", f"+{float(stats.total_profit):,.2f}", "所有盈利交易累计"),
            ("总亏损", f"-{float(stats.total_loss):,.2f}", "所有亏损交易累计"),
            ("净利润", f"{float(stats.net_profit):,.2f}", "总盈利-总亏损"),
            ("平均盈利", f"+{float(stats.avg_profit):,.2f}", "单笔盈利交易平均"),
            ("平均亏损", f"-{float(stats.avg_loss):,.2f}", "单笔亏损交易平均"),
        ]
        row = self._add_table(ws, row, ["项目", "金额", "备注"], profit_data)

        # 持仓时间统计
        row = self._add_section_header(ws, row + 1, "三、持仓时间分析")
        holding_data = [
            ("平均持仓天数", f"{stats.avg_holding_days:.1f}天", "所有交易平均"),
            (
                "盈利交易平均持仓",
                f"{stats.avg_winning_holding_days:.1f}天",
                "持有时间较长",
            ),
            (
                "亏损交易平均持仓",
                f"{stats.avg_losing_holding_days:.1f}天",
                "持有时间较短",
            ),
            ("最长持仓", f"{stats.max_holding_days}天", ""),
            ("最短持仓", f"{stats.min_holding_days}天", ""),
        ]
        row = self._add_table(ws, row, ["指标", "天数", "说明"], holding_data)

        # 市场分布
        row = self._add_section_header(ws, row + 1, "四、市场分布")
        market_headers = ["市场", "交易笔数", "胜率", "净盈亏"]
        market_data = []
        for market, ms in stats.market_stats.items():
            market_name = {"HK": "港股", "US": "美股", "SH": "沪市", "SZ": "深市"}.get(
                market, market
            )
            market_data.append(
                (
                    market_name,
                    f"{ms.total_trades}笔",
                    f"{ms.win_rate:.1%}",
                    f"{float(ms.net_profit):,.2f}",
                )
            )
        if market_data:
            row = self._add_table(ws, row, market_headers, market_data)

        # 期权统计（如果有）
        if stats.option_total_trades > 0:
            row = self._add_section_header(ws, row + 1, "五、期权交易统计")
            option_data = [
                ("期权交易笔数", f"{stats.option_total_trades}笔", ""),
                ("期权胜率", f"{stats.option_win_rate:.1%}", ""),
                ("期权净盈亏", f"{float(stats.option_net_profit):,.2f}", ""),
            ]
            row = self._add_table(ws, row, ["指标", "数值", "备注"], option_data)

        # 设置列宽
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 30

    def _add_section_header(self, ws, row: int, title: str) -> int:
        """添加章节标题"""
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
        return row + 1

    def _add_table(
        self, ws, start_row: int, headers: list[str], data: list[tuple]
    ) -> int:
        """添加数据表格"""
        # 表头
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center")

        # 数据行
        for row_idx, row_data in enumerate(data, start=start_row + 1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                cell.alignment = Alignment(
                    horizontal="left" if col_idx == 1 else "center"
                )

                # 盈亏颜色
                if isinstance(value, str):
                    if value.startswith("+"):
                        cell.font = self.PROFIT_FONT
                    elif value.startswith("-"):
                        cell.font = self.LOSS_FONT

        return start_row + len(data) + 1
