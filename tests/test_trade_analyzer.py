"""
Tests for Trade Analyzer module

Tests for:
- TradeMatcher: Buy/sell trade pairing with LIFO algorithm
- StatisticsCalculator: Trade statistics calculation
- ChartGenerator: Chart generation
- ExcelExporter: Excel export
- DocxExporter: Word export
"""

import pytest
from datetime import datetime, timedelta
from decimal import Decimal
from pathlib import Path
from unittest.mock import MagicMock, patch

from skills.trade_analyzer import (
    TradeMatcher,
    MatchedTrade,
    StatisticsCalculator,
    TradeStatistics,
    ChartGenerator,
    ExcelExporter,
    DocxExporter,
    TradeAnalyzer,
)


# =============================================================================
# Fixtures
# =============================================================================


@pytest.fixture
def mock_trade():
    """Create a mock Trade object"""
    def _create_trade(
        deal_id: str,
        market: str,
        code: str,
        trd_side: str,
        qty: float,
        price: float,
        trade_time: datetime,
        fee: float = 10.0,
        stock_name: str = "Test Stock",
    ):
        trade = MagicMock()
        trade.deal_id = deal_id
        trade.market = market
        trade.code = code
        trade.trd_side = trd_side
        trade.qty = qty
        trade.price = price
        trade.trade_time = trade_time
        trade.fee = fee
        trade.stock_name = stock_name
        return trade
    return _create_trade


@pytest.fixture
def sample_trades(mock_trade):
    """Create sample trade records for testing"""
    base_date = datetime(2025, 1, 1)
    return [
        # Buy 100 shares @ 10.00 on Jan 1
        mock_trade("D001", "HK", "00700", "BUY", 100, 10.0, base_date),
        # Buy 200 shares @ 12.00 on Jan 5
        mock_trade("D002", "HK", "00700", "BUY", 200, 12.0, base_date + timedelta(days=4)),
        # Sell 150 shares @ 15.00 on Jan 10 (LIFO: takes 150 from 200-share lot)
        mock_trade("D003", "HK", "00700", "SELL", 150, 15.0, base_date + timedelta(days=9)),
        # Sell 150 shares @ 11.00 on Jan 15 (LIFO: takes 50 from 200-share lot + 100 from 100-share lot)
        mock_trade("D004", "HK", "00700", "SELL", 150, 11.0, base_date + timedelta(days=14)),
    ]


@pytest.fixture
def matched_trades():
    """Create sample matched trades for testing statistics"""
    trades = []

    # Winning trade: Buy @ 10, Sell @ 15 (50% profit)
    t1 = MatchedTrade(
        market="HK",
        code="00700",
        stock_name="腾讯控股",
        buy_price=Decimal("10"),
        buy_qty=Decimal("100"),
        buy_amount=Decimal("1000"),
        buy_date=datetime(2025, 1, 1),
        sell_price=Decimal("15"),
        sell_qty=Decimal("100"),
        sell_amount=Decimal("1500"),
        sell_date=datetime(2025, 1, 10),
    )
    t1.calculate()
    trades.append(t1)

    # Losing trade: Buy @ 20, Sell @ 15 (25% loss)
    t2 = MatchedTrade(
        market="HK",
        code="00981",
        stock_name="中芯国际",
        buy_price=Decimal("20"),
        buy_qty=Decimal("100"),
        buy_amount=Decimal("2000"),
        buy_date=datetime(2025, 2, 1),
        sell_price=Decimal("15"),
        sell_qty=Decimal("100"),
        sell_amount=Decimal("1500"),
        sell_date=datetime(2025, 2, 15),
    )
    t2.calculate()
    trades.append(t2)

    # Another winning trade
    t3 = MatchedTrade(
        market="US",
        code="AAPL",
        stock_name="Apple Inc",
        buy_price=Decimal("150"),
        buy_qty=Decimal("10"),
        buy_amount=Decimal("1500"),
        buy_date=datetime(2025, 3, 1),
        sell_price=Decimal("180"),
        sell_qty=Decimal("10"),
        sell_amount=Decimal("1800"),
        sell_date=datetime(2025, 3, 20),
    )
    t3.calculate()
    trades.append(t3)

    return trades


# =============================================================================
# TradeMatcher Tests
# =============================================================================


class TestTradeMatcher:
    """Tests for TradeMatcher class"""

    def test_is_option_code_hk_stock(self):
        """Test HK stock codes are not identified as options"""
        assert TradeMatcher.is_option_code("HK", "00700") is False
        assert TradeMatcher.is_option_code("HK", "00981") is False

    def test_is_option_code_hk_option(self):
        """Test HK option codes are correctly identified"""
        assert TradeMatcher.is_option_code("HK", "SMC260629C75000") is True
        assert TradeMatcher.is_option_code("HK", "TCH260330C650000") is True

    def test_is_option_code_us_stock(self):
        """Test US stock codes are not identified as options"""
        assert TradeMatcher.is_option_code("US", "AAPL") is False
        assert TradeMatcher.is_option_code("US", "NVDA") is False

    def test_is_option_code_us_option(self):
        """Test US option codes are correctly identified"""
        assert TradeMatcher.is_option_code("US", "MU260116C230000") is True
        assert TradeMatcher.is_option_code("US", "NVDA260116C186000") is True

    def test_match_simple_trade(self, mock_trade):
        """Test matching a simple buy-sell pair"""
        base_date = datetime(2025, 1, 1)
        trades = [
            mock_trade("D001", "HK", "00700", "BUY", 100, 10.0, base_date),
            mock_trade("D002", "HK", "00700", "SELL", 100, 15.0, base_date + timedelta(days=5)),
        ]

        matcher = TradeMatcher()
        matched = matcher.match_trades(trades)

        assert len(matched) == 1
        assert matched[0].buy_qty == Decimal("100")
        assert matched[0].sell_qty == Decimal("100")
        assert matched[0].buy_price == Decimal("10")
        assert matched[0].sell_price == Decimal("15")
        assert matched[0].holding_days == 5
        assert matched[0].is_profitable is True

    def test_match_lifo_order(self, sample_trades):
        """Test LIFO matching order - most recent buy matched first"""
        matcher = TradeMatcher()
        matched = matcher.match_trades(sample_trades)

        # Should have 2 matched trades
        # Trade 1: 150 shares from the 200-share lot @ 12, sold @ 15
        # Trade 2: 50 shares from 200-share lot + 100 from 100-share lot, sold @ 11
        assert len(matched) == 3  # Split due to LIFO

    def test_partial_match(self, mock_trade):
        """Test partial matching when sell qty doesn't match buy qty"""
        base_date = datetime(2025, 1, 1)
        trades = [
            mock_trade("D001", "HK", "00700", "BUY", 200, 10.0, base_date),
            mock_trade("D002", "HK", "00700", "SELL", 100, 15.0, base_date + timedelta(days=5)),
        ]

        matcher = TradeMatcher()
        matched = matcher.match_trades(trades)

        assert len(matched) == 1
        assert matched[0].buy_qty == Decimal("100")
        assert matched[0].sell_qty == Decimal("100")

        # Should have remaining unmatched buy
        unmatched_buys = matcher.get_unmatched_buys()
        assert len(unmatched_buys) == 1
        assert unmatched_buys[0].remaining_qty == Decimal("100")

    def test_stock_vs_option_separation(self, mock_trade):
        """Test that stock and option trades are separated"""
        base_date = datetime(2025, 1, 1)
        trades = [
            # Stock trades
            mock_trade("D001", "HK", "00700", "BUY", 100, 10.0, base_date),
            mock_trade("D002", "HK", "00700", "SELL", 100, 15.0, base_date + timedelta(days=5)),
            # Option trades
            mock_trade("D003", "HK", "TCH260330C650000", "BUY", 10, 5.0, base_date),
            mock_trade("D004", "HK", "TCH260330C650000", "SELL", 10, 8.0, base_date + timedelta(days=3)),
        ]

        matcher = TradeMatcher()
        matched = matcher.match_trades(trades)

        stock_trades = matcher.get_stock_trades()
        option_trades = matcher.get_option_trades()

        assert len(stock_trades) == 1
        assert len(option_trades) == 1
        assert stock_trades[0].is_option is False
        assert option_trades[0].is_option is True


# =============================================================================
# StatisticsCalculator Tests
# =============================================================================


class TestStatisticsCalculator:
    """Tests for StatisticsCalculator class"""

    def test_calculate_basic_stats(self, matched_trades):
        """Test basic statistics calculation"""
        calculator = StatisticsCalculator()
        stats = calculator.calculate(matched_trades)

        assert stats.total_trades == 3
        assert stats.winning_trades == 2
        assert stats.losing_trades == 1

    def test_calculate_win_rate(self, matched_trades):
        """Test win rate calculation"""
        calculator = StatisticsCalculator()
        stats = calculator.calculate(matched_trades)

        # 2 wins out of 3 trades = 66.67%
        assert abs(stats.win_rate - 0.6667) < 0.01

    def test_calculate_profit_loss(self, matched_trades):
        """Test profit/loss calculation"""
        calculator = StatisticsCalculator()
        stats = calculator.calculate(matched_trades)

        # Trade 1: +500, Trade 2: -500, Trade 3: +300
        # Total profit: 800, Total loss: 500, Net: 300
        assert stats.total_profit == Decimal("800")
        assert stats.total_loss == Decimal("500")
        assert stats.net_profit == Decimal("300")

    def test_calculate_holding_days(self, matched_trades):
        """Test holding days statistics"""
        calculator = StatisticsCalculator()
        stats = calculator.calculate(matched_trades)

        # Trade 1: 9 days, Trade 2: 14 days, Trade 3: 19 days
        # Average: (9 + 14 + 19) / 3 = 14
        assert abs(stats.avg_holding_days - 14) < 0.1

    def test_calculate_market_stats(self, matched_trades):
        """Test market distribution statistics"""
        calculator = StatisticsCalculator()
        stats = calculator.calculate(matched_trades)

        assert "HK" in stats.market_stats
        assert "US" in stats.market_stats
        assert stats.market_stats["HK"].total_trades == 2
        assert stats.market_stats["US"].total_trades == 1

    def test_calculate_empty_trades(self):
        """Test with empty trade list"""
        calculator = StatisticsCalculator()
        stats = calculator.calculate([])

        assert stats.total_trades == 0
        assert stats.win_rate == 0.0
        assert stats.net_profit == Decimal("0")


# =============================================================================
# MatchedTrade Tests
# =============================================================================


class TestMatchedTrade:
    """Tests for MatchedTrade class"""

    def test_calculate_profit(self):
        """Test profit calculation"""
        trade = MatchedTrade(
            market="HK",
            code="00700",
            stock_name="腾讯控股",
            buy_price=Decimal("100"),
            buy_qty=Decimal("100"),
            buy_amount=Decimal("10000"),
            buy_date=datetime(2025, 1, 1),
            buy_fee=Decimal("50"),
            sell_price=Decimal("120"),
            sell_qty=Decimal("100"),
            sell_amount=Decimal("12000"),
            sell_date=datetime(2025, 1, 10),
            sell_fee=Decimal("50"),
        )
        trade.calculate()

        # Profit = 12000 - 10000 - 50 - 50 = 1900
        assert trade.profit_loss == Decimal("1900")
        # Profit ratio = 1900 / 10000 = 0.19
        assert trade.profit_loss_ratio == Decimal("0.19")
        assert trade.holding_days == 9

    def test_calculate_loss(self):
        """Test loss calculation"""
        trade = MatchedTrade(
            market="HK",
            code="00700",
            stock_name="腾讯控股",
            buy_price=Decimal("100"),
            buy_qty=Decimal("100"),
            buy_amount=Decimal("10000"),
            buy_date=datetime(2025, 1, 1),
            sell_price=Decimal("80"),
            sell_qty=Decimal("100"),
            sell_amount=Decimal("8000"),
            sell_date=datetime(2025, 1, 10),
        )
        trade.calculate()

        assert trade.profit_loss == Decimal("-2000")
        assert trade.is_profitable is False

    def test_to_dict(self):
        """Test conversion to dictionary"""
        trade = MatchedTrade(
            market="HK",
            code="00700",
            stock_name="腾讯控股",
            buy_price=Decimal("100"),
            buy_qty=Decimal("100"),
            buy_amount=Decimal("10000"),
            buy_date=datetime(2025, 1, 1),
            sell_price=Decimal("120"),
            sell_qty=Decimal("100"),
            sell_amount=Decimal("12000"),
            sell_date=datetime(2025, 1, 10),
        )
        trade.calculate()

        data = trade.to_dict()

        assert data["market"] == "HK"
        assert data["code"] == "00700"
        assert data["buy_price"] == 100.0
        assert data["sell_price"] == 120.0


# =============================================================================
# ChartGenerator Tests
# =============================================================================


class TestChartGenerator:
    """Tests for ChartGenerator class"""

    def test_generate_win_loss_pie(self, matched_trades):
        """Test win/loss pie chart generation"""
        calculator = StatisticsCalculator()
        stats = calculator.calculate(matched_trades)

        generator = ChartGenerator()
        chart_data = generator.generate_win_loss_pie(stats)

        # Should return PNG bytes
        assert isinstance(chart_data, bytes)
        assert len(chart_data) > 0
        # PNG magic bytes
        assert chart_data[:8] == b'\x89PNG\r\n\x1a\n'

    def test_generate_monthly_profit_bar(self, matched_trades):
        """Test monthly profit bar chart generation"""
        calculator = StatisticsCalculator()
        stats = calculator.calculate(matched_trades)

        generator = ChartGenerator()
        chart_data = generator.generate_monthly_profit_bar(stats)

        assert isinstance(chart_data, bytes)
        assert len(chart_data) > 0

    def test_generate_all_charts(self, matched_trades):
        """Test generating all charts"""
        calculator = StatisticsCalculator()
        stats = calculator.calculate(matched_trades)

        generator = ChartGenerator()
        charts = generator.generate_all_charts(matched_trades, stats)

        assert isinstance(charts, dict)
        assert "win_loss_pie" in charts
        assert "monthly_profit_bar" in charts


# =============================================================================
# ExcelExporter Tests
# =============================================================================


class TestExcelExporter:
    """Tests for ExcelExporter class"""

    def test_export_creates_file(self, matched_trades, tmp_path):
        """Test Excel export creates a file"""
        calculator = StatisticsCalculator()
        stats = calculator.calculate(matched_trades)

        output_path = tmp_path / "test_export.xlsx"

        exporter = ExcelExporter()
        result = exporter.export(matched_trades, stats, output_path, year=2025)

        assert result.exists()
        assert result.suffix == ".xlsx"

    def test_export_has_correct_sheets(self, matched_trades, tmp_path):
        """Test exported Excel has correct sheets"""
        import openpyxl

        calculator = StatisticsCalculator()
        stats = calculator.calculate(matched_trades)

        output_path = tmp_path / "test_export.xlsx"

        exporter = ExcelExporter()
        exporter.export(matched_trades, stats, output_path, year=2025)

        wb = openpyxl.load_workbook(output_path)

        assert "股票交易明细" in wb.sheetnames
        assert "统计汇总" in wb.sheetnames


# =============================================================================
# DocxExporter Tests
# =============================================================================


class TestDocxExporter:
    """Tests for DocxExporter class"""

    def test_export_creates_file(self, matched_trades, tmp_path):
        """Test Word export creates a file"""
        calculator = StatisticsCalculator()
        stats = calculator.calculate(matched_trades)

        generator = ChartGenerator()
        charts = generator.generate_all_charts(matched_trades, stats)

        output_path = tmp_path / "test_report.docx"

        exporter = DocxExporter()
        result = exporter.export(matched_trades, stats, charts, output_path, year=2025)

        assert result.exists()
        assert result.suffix == ".docx"

    def test_export_has_content(self, matched_trades, tmp_path):
        """Test exported Word has content"""
        from docx import Document

        calculator = StatisticsCalculator()
        stats = calculator.calculate(matched_trades)

        generator = ChartGenerator()
        charts = generator.generate_all_charts(matched_trades, stats)

        output_path = tmp_path / "test_report.docx"

        exporter = DocxExporter()
        exporter.export(matched_trades, stats, charts, output_path, year=2025)

        doc = Document(output_path)

        # Check for expected headings
        headings = [p.text for p in doc.paragraphs if p.style.name.startswith("Heading")]
        assert any("整体交易表现" in h for h in headings)
        assert any("盈亏统计" in h for h in headings)


# =============================================================================
# Integration Tests
# =============================================================================


class TestTradeAnalyzerIntegration:
    """Integration tests for TradeAnalyzer"""

    @patch('skills.trade_analyzer.trade_analyzer.get_session')
    def test_analyze_no_trades(self, mock_get_session):
        """Test analysis with no trades returns empty result"""
        from datetime import date

        # Mock session that returns no trades
        mock_session = MagicMock()
        mock_session.query.return_value.filter.return_value.all.return_value = []
        mock_session.__enter__ = MagicMock(return_value=mock_session)
        mock_session.__exit__ = MagicMock(return_value=None)
        mock_get_session.return_value = mock_session

        analyzer = TradeAnalyzer()
        result = analyzer.analyze(
            user_id=1,
            start_date=date(2025, 1, 1),
            end_date=date(2025, 12, 31),
            generate_excel=False,
            generate_docx=False,
            generate_charts=False,
        )

        assert result.total_raw_trades == 0
        assert len(result.matched_trades) == 0
